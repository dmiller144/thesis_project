# -*- coding: utf-8 -*-
"""
Created on Mon Dec 21 13:54:18 2020
This code returns the pumping and storage configuration which has the lowest energy requirements for a given network.

Note: The network must have 1 pump, 1 reservoir and no pump scheduling controls. 
The network must ONLY include pipe diameters listed below. 
@author: dmi002
"""

import wntr
# import numpy as np
import pandas as pd
import math
# import xlsxwriter
from openpyxl import Workbook
from openpyxl import load_workbook
import os
# import matplotlib as plt

# Pipe Cost Table for T/L
pipecost_dict = {"Pipe_Diameter\n(mm)": [63.0, 75.0, 90.0, 110.0, 125.0, 140.0, 160.0, 200.0, 225.0, 250.0, 315.0, 355.0, 400.0, 450.0, 500.0, 600.0, 700.0, 800.0],
           "Pipe_Material": ['PVC', 'PVC', 'PVC', 'PVC', 'PVC', 'PVC', 'PVC', 'PVC', 'PVC', 'PVC', 'PVC', 'PVC', 'Ductile Iron', 'Ductile Iron', 'Ductile Iron', 'Ductile Iron', 'Ductile Iron', 'Ductile Iron'],
           "Supply_and_Installed_Cost\n($ USD/m)": [14.0, 17.0, 21.0, 28.0, 34.0, 41.0, 52.0, 79.0, 100.0, 120.0, 190.0, 240.0, 253.0, 295.0, 348.0, 462.0, 590.0, 732.0]
           }
pipecost_df = pd.DataFrame.from_dict(pipecost_dict)

pipe_diameters = pipecost_df["Pipe_Diameter\n(mm)"].tolist()
pipe_roughness = 0.5

# Global inputs
min_p_req = 20
max_headloss = 20 # m/km
energy_price = 0.42 # € /kWh

# Annual O&M as % of investment
OM_pipes = 0.5
OM_pumps = 2.0
OM_tanks = 0.8

# Loan conditions
Repay = 20.0 # years
Interest = 6.0 # annual interest rate on %

# Load model
INP_file_name = 'Dili_half' # input('Enter the filename of the network you wish to process: ')
inp_file = f'C:/Users/dmi002/Desktop/Python WIP/thesis_project/{INP_file_name}.inp'
wn = wntr.network.WaterNetworkModel(inp_file)

# Creating Directory for results and networks
try:
    os.makedirs(f'C:/Users/dmi002/Desktop/Python WIP/thesis_project/{INP_file_name}_results')
    os.makedirs(f'C:/Users/dmi002/Desktop/Python WIP/thesis_project/{INP_file_name}_networks')
except:
    pass

# Original Pipe diameters
pipe_info = {}
pipe_names = [wn.get_link(pipe).name for pipe in wn.pipe_name_list]
original_diameter = [wn.get_link(pipe).diameter * 1000 for pipe in wn.pipe_name_list]
pipe_length = [wn.get_link(pipe).length for pipe in wn.pipe_name_list]

tank_con_length = 1
tank_con_diameter = 1
pipe_names.append('tank_connection')
original_diameter.append(None)
pipe_length.append(tank_con_length)                  
pipe_info = {'Pipe_name': pipe_names,'Original_Diameter': original_diameter, 'Length': pipe_length}
pipe_df = pd.DataFrame.from_dict(pipe_info)

# Setting Water Quality Option
wn.options.quality.parameter = 'AGE'
original_duration = wn.options.time.duration

# Setting global energy options
wn.options.energy.global_efficiency = 75.0
wn.options.energy.global_price = energy_price/(3.6*10**6) #converting € 0.15 per kWh into €  per joules 
wn.options.energy.global_pattern = None 

reservoirs = wn.reservoir_name_list
junctions = wn.junction_name_list

# Running Sim
wn.options.time.duration = 7 * 24 * 3600
sim = wntr.sim.EpanetSimulator(wn)
results = sim.run_sim()

# Water Age
water_age = (results.node['quality'].loc[:, junctions].max()/3600).to_list()
water_age_rounded = [round(elem,2) for elem in water_age]
wn.options.time.duration = original_duration

# Calculating required balancing tank volume
avg_flow = wntr.metrics.hydraulic.average_expected_demand(wn).sum()
demand = wntr.metrics.hydraulic.expected_demand(wn, end_time=wn.options.time.duration - wn.options.time.hydraulic_timestep).sum(axis=1)
cumulative = (avg_flow - demand).cumsum()
tank_volume = (cumulative.max() + abs(cumulative.min())) * wn.options.time.hydraulic_timestep

# Establishing fixed tank parameters
tank_min_lvl = 2.0
tank_max_lvl = 6.0
tank_diameter = math.sqrt(4/math.pi*tank_volume/(tank_max_lvl - tank_min_lvl))
tank_init_lvl = (tank_max_lvl - tank_min_lvl)/2 + tank_min_lvl
tank_fin_lvl = tank_max_lvl

# Setting Duty_Flow to avg_flow
pump = wn.get_link(wn.pump_name_list[0])
pump_end_node = str(pump.end_node)
Duty_Head = pump.get_head_curve_coefficients()[0] * 3/4
Duty_Flow = avg_flow
curve = pump.pump_curve_name
wn.get_curve(curve).points = [(Duty_Flow, Duty_Head)]
sim = wntr.sim.EpanetSimulator(wn)
results = sim.run_sim()

min_p = results.node['pressure'].min().min()

# Adding pump variable speed pattern
multipliers = [1]
wn.add_pattern('VarSpeed', multipliers)

counter = 0

# Creating Workbook
wb = Workbook()
wb.active.title = "Summary"
wb.active['A1'] = 'Summary of results for tank connected to each node'
for j in junctions:
    ws = wb.create_sheet(j)
    ws.title = j
wb.save(f'{INP_file_name}_results/{INP_file_name}_results.xlsx')

# Lists to store results
(node_list, Duty_Head_list, Duty_Flow_list, ActAvgFlow_list, Cost_list, Energy_list, Tank_elev_list,
Tank_height_list, Tank_volume_list, Min_p_list, Critical_hr_list, Critical_node_list, Critical_pipes_list, 
Pump_inv_cost_list, Tank_inv_cost_list, Pipe_replacement_cost_list, Total_inv_cost_list, Maintenance_list,
Annuity_list, Total_annual_exp_list) = ([] for i in range(20))

junction = ['a01', 'a02'] # This line replaces junctions for testing the code
# wn.write_inpfile(f'{INP_file_name} test tank_con_node {j}.inp', version=2.2)
# j = 'R_TIBAR_1'
for j in junctions:
    # Resetting Pipe diameters to original
    pipe_df.set_index('Pipe_name', inplace=True)
    for pipe_name in wn.pipe_name_list:
        # pipe_name = 'p03'
        pipe = wn.get_link(pipe_name)
        pipe.diameter = pipe_df['Original_Diameter'].loc[pipe_name] / 1000
    pipe_df = pipe_df.reset_index()
    
    
    
    # Adding Balancing Tank
    wn.add_tank('Balancing_tank', elevation=wn.get_node(j).elevation, init_level=tank_init_lvl, min_level=0, 
                max_level=10000, diameter=tank_diameter, coordinates=(wn.get_node(j).coordinates[0] + 30, wn.get_node(j).coordinates[1] + 30))
    wn.add_pipe('tank_connection', 'Balancing_tank', j, length=tank_con_length, diameter=tank_con_diameter, roughness=0.5)
    
    # Initialising parameters
    tank_init_lvl = (tank_max_lvl - tank_min_lvl)/2 + tank_min_lvl
    Duty_Head = pump.get_head_curve_coefficients()[0] * 3/4
    Duty_Flow = avg_flow
    curve = pump.pump_curve_name
    wn.get_curve(curve).points = [(Duty_Flow, Duty_Head)]
    multipliers = [1]
    
    sim = wntr.sim.EpanetSimulator(wn)
    results = sim.run_sim()
    tank_fin_lvl = results.node['pressure'].loc[24*3600, 'Balancing_tank']
    min_p = results.node['pressure'].loc[:, junctions].min().min()
    min_pumped_flow = results.link['flowrate'].loc[:, wn.pump_name_list[0]].min()
    
    # Critical Pipes
    df = results.link['headloss'].loc[:, wn.pipe_name_list]
    min_p = results.node['pressure'].loc[:, junctions].min().min()
    critical_pipes_times = df[df > max_headloss].dropna(axis=1, how='all').dropna(how='all')
    critical_pipes = critical_pipes_times.columns.to_list()
    # wn.write_inpfile(f'{INP_file_name} test2 tank_con_node {j}.inp', version=2.2)
    # [0.025, 0.05, 0.08, 0.1, 0.125, 0.15, 0.2, 0.25, 0.3, 0.4, 0.5, 0.6]
    # diam = wn.get_link('p03').diameter
    while len(critical_pipes) >= 1:
        for pipe_name in critical_pipes:
            # pipe_name = 'p03'
            pipe = wn.get_link(pipe_name)
            indexer = min(range(len(pipe_diameters)), key=lambda i: abs(pipe_diameters[i] - (pipe.diameter * 1000)))
            if pipe_diameters[indexer] <= (pipe.diameter * 1000):
                if indexer == len(pipe_diameters)-1:
                    indexer = indexer
                else:
                    indexer +=1
            pipe.diameter = pipe_diameters[indexer] / 1000
            # print(pipe_name, 'is now diameter:', pipe.diameter, 'm')    
            pipe.roughness = pipe_roughness  
        sim = wntr.sim.EpanetSimulator(wn)
        results = sim.run_sim()
        # Critical Pipes
        df = results.link['headloss'].loc[:, wn.pipe_name_list]
        min_p = results.node['pressure'].loc[:, junctions].min().min()
        critical_pipes_times = df[df > max_headloss].dropna(axis=1, how='all').dropna(how='all')
        critical_pipes = critical_pipes_times.columns.to_list()
    
    # wn.write_inpfile(f'{INP_file_name} test tank_con_node {j}.inp', version=2.2)
    
    # This loop sets the pump head such that the min P in network equals the desired P
    while abs(min_p - min_p_req) > 0.1:
        while abs(min_p - min_p_req) > 0.1:
            
            # This loops ensures tank balancing
            while abs(tank_init_lvl - tank_fin_lvl) > 0.01:
                wn.get_node('Balancing_tank').init_level = tank_fin_lvl
                sim = wntr.sim.EpanetSimulator(wn)
                results = sim.run_sim()
                tank_fin_lvl = results.node['pressure'].loc[24*3600, 'Balancing_tank']
                tank_init_lvl = wn.get_node('Balancing_tank').init_level
            Duty_Head = Duty_Head + min_p_req - min_p + 0.1
            wn.get_curve(curve).points = [(Duty_Flow, Duty_Head)]
                    
            # the head has to be ablet to provide the avg flow at all times
            # print('Duty Head:', Duty_Head, 'tank_init_lvl:', tank_init_lvl)
            sim = wntr.sim.EpanetSimulator(wn)
            results = sim.run_sim()
            tank_fin_lvl = results.node['pressure'].loc[24*3600, 'Balancing_tank']
            tank_init_lvl = wn.get_node('Balancing_tank').init_level
            
            # This loops ensures tank balancing
            while abs(tank_init_lvl - tank_fin_lvl) > 0.01:
                wn.get_node('Balancing_tank').init_level = tank_fin_lvl
                sim = wntr.sim.EpanetSimulator(wn)
                results = sim.run_sim()
                tank_fin_lvl = results.node['pressure'].loc[24*3600, 'Balancing_tank']
                tank_init_lvl = wn.get_node('Balancing_tank').init_level
            min_p = results.node['pressure'].loc[:, junctions].min().min()
        
        pump_flows = results.link['flowrate'].loc[:, wn.pump_name_list[0]]*1000
        
        # This loop applies the flow balancing method is the standard deviation of the pump flows is greater than 1
        while pump_flows.std() > 1:
            """Start code to pump avg flow"""
            
            # Adding flow control valve to pump avg flow
            wn.add_junction('Dummy_node', elevation=wn.get_node(pump_end_node).elevation, coordinates=(wn.get_node(pump_end_node).coordinates[0]-30, wn.get_node(pump_end_node).coordinates[1]-30))
            wn.add_valve('Dummy_FCV','Dummy_node', pump_end_node, diameter=1, valve_type='FCV', minor_loss=0, setting=avg_flow)
            wn.remove_link(wn.pump_name_list[0])
            wn.add_pump('Pump', wn.reservoir_name_list[0], 'Dummy_node', pump_type='HEAD', pump_parameter=curve, speed=1.0, pattern='VarSpeed')
            
            sim = wntr.sim.EpanetSimulator(wn)
            results = sim.run_sim()
                
            # Setting Duty Head 500mwc above max pressure to ensure Qavg is maintained
            Duty_Head = results.node['pressure'].loc[:, pump_end_node].max()+500
            wn.get_curve(curve).points = [(Duty_Flow, Duty_Head)]
            
            sim = wntr.sim.EpanetSimulator(wn)
            results = sim.run_sim()
            
            # Range of Heads the pump sees
            pumped_head = results.node['pressure'].loc[:, pump_end_node]
            Duty_Head = pumped_head.mean()
            wn.get_curve(curve).points = [(Duty_Flow, Duty_Head)]
                
            #Creating pattern of pump speeds
            multipliers = (0.25+3*pumped_head.array/(4*(Duty_Head)))**0.5
            multipliers = multipliers[:-1]
            VarSpeed = wn.get_pattern('VarSpeed')
            VarSpeed.multipliers = multipliers
                
            # Removing FCV
            wn.remove_link('Pump')
            wn.remove_link('Dummy_FCV')
            wn.remove_node('Dummy_node')
            wn.add_pump('Pump', wn.reservoir_name_list[0], pump_end_node, pump_type='HEAD', pump_parameter=curve, speed=1.0, pattern='VarSpeed')
            
            """End code to pump avg flow"""
            
            sim = wntr.sim.EpanetSimulator(wn)
            results = sim.run_sim()
            pump_flows = results.link['flowrate'].loc[:, wn.pump_name_list[0]]*1000
            min_p = results.node['pressure'].loc[:, junctions].min().min()
                    
    sim = wntr.sim.EpanetSimulator(wn)
    results = sim.run_sim()
    min_p = results.node['pressure'].loc[:, junctions].min().min()
    counter += 1  # print(counter) 
    
    # Adjusting tank elevation up
    wn.get_node('Balancing_tank').elevation = wn.get_node('Balancing_tank').elevation + results.node['pressure'].loc[:, 'Balancing_tank'].min() - tank_min_lvl
    wn.get_node('Balancing_tank').init_level = wn.get_node('Balancing_tank').init_level - results.node['pressure'].loc[:, 'Balancing_tank'].min() + tank_min_lvl
    tank_elev = wn.get_node('Balancing_tank').elevation
    tank_init_lvl = wn.get_node('Balancing_tank').init_level
    tank_height = tank_elev - wn.get_node(j).elevation
           
    # Energy Calculations
    pump_flowrate = results.link['flowrate'].loc[:, wn.pump_name_list]
    head = results.node['head']
    pump_energy = wntr.metrics.pump_energy(pump_flowrate, head, wn)
    pump_cost = wntr.metrics.pump_cost(pump_flowrate, head, wn)
       
    # Critical Pipes
    sim = wntr.sim.EpanetSimulator(wn)
    results = sim.run_sim()
    df = results.link['headloss'].loc[:, wn.pipe_name_list]
    min_p = results.node['pressure'].loc[:, junctions].min().min()
    critical_pipes_times = df[df > max_headloss].dropna(axis=1, how='all').dropna(how='all')
    critical_pipes = critical_pipes_times.columns.to_list()
    
    # Adding new pipe info and calculating replacement costs
    new_diameter = [wn.get_link(pipe).diameter * 1000 for pipe in wn.pipe_name_list]
    
    
    pipe_df["New_Diameter\n(mm)"] = new_diameter
    pipe_df = pipe_df.merge(pipecost_df, left_on="New_Diameter\n(mm)", right_on="Pipe_Diameter\n(mm)", how='left')
    # pipe_df = pipe_df.sort_values(by="New_Diameter\n(mm)")
    # pipecost_df = pipecost_df.sort_values(by="Pipe_Diameter\n(mm)")
    # pipe_df = pd.merge_asof(pipe_df, pipecost_df, left_on="New_Diameter\n(mm)", right_on="Pipe_Diameter\n(mm)", direction="nearest")
    # print(type(pipe_df['Original_Diameter'].iloc[5]), '\n', type(pipe_df["New_Diameter\n(mm)"].iloc[5]))
    
    pipe_df['Replacement_Cost'] = pipe_df['Length'] * pipe_df["Supply_and_Installed_Cost\n($ USD/m)"].where(pipe_df['Original_Diameter'] != pipe_df["New_Diameter\n(mm)"])
    pipe_df['Total_Cost'] = pipe_df['Length'] * pipe_df["Supply_and_Installed_Cost\n($ USD/m)"]
    pipe_df.to_clipboard()
    pipe_df = pipe_df.drop("Pipe_Diameter\n(mm)", 1)
    pipe_df = pipe_df.drop("Pipe_Material", 1)
    pipe_df.set_index('Pipe_name', inplace=True)
    
    # Tank Cost calculation
    tank_inv_cost = 300_000 + 150 * tank_volume + 3 * tank_height * tank_volume
    
    # Pump Cost calculation
    pump_inv_cost = 5000 * ((pump.get_design_flow()*1.8*3600)**0.8)
    
    # Total Investment cost
    total_inv_cost = tank_inv_cost + pump_inv_cost + pipe_df['Replacement_Cost'].sum()
    
    #Maintenance Cost Calculation
    maintenance = (OM_pipes/100)*pipe_df['Total_Cost'].sum() + (OM_pumps/100)*pump_inv_cost + (OM_tanks/100)*tank_inv_cost
    pipe_df = pipe_df.drop("Total_Cost", 1)
    # Annuity Calculation
    Annuity = (Interest/100*(1+Interest/100)**Repay)/((1+Interest/100)**Repay-1)*total_inv_cost
    
    # Total Annual Expenditure Cost
    Total_annual_exp = maintenance + pump_cost[0:24].sum()[0]*3600*365 + Annuity
    
    
    # Outputs to console
    print('\nProgress:', counter, '/', len(junctions), 'Nodes Processed')
    print(f'Summary of pumping energy for tank connected to node {j}:')
    print('Pump Parameters: Duty Flow:', round(Duty_Flow*1000, 2),'L/s',  'Duty Head:', round(Duty_Head, 2), 'm', f'Pump Cost: € {round(pump_inv_cost,2):,}')
    print('Actual average pumped flow:', round(pump_flows.mean(), 2), 'L/s')
    print('Cost € ', round(pump_cost[0:24].sum()[0]*3600, 2), 'per day')
    print('Energy:', f'{round(pump_energy[0:24].sum()[0]/1000, 2):,}', 'kWh/day')
    print('Minimum Pressure:', round(min_p, 2), 'mwc','at time:', results.node['pressure'].loc[:, str(results.node['pressure'].loc[:, junctions].min().idxmin())].idxmin()/3600 ,
          'At node:', results.node['pressure'].loc[:, junctions].min().idxmin())
    print('Tank Elevation:', round(tank_elev, 2), 'm', 'Tank Height above ground:', round(tank_height, 2), 'm', 'Tank Volume:', f'{round(tank_volume, 2):,}', 'm3', 'Tank Cost: € ', f'{round(tank_inv_cost, 2):,}')
    print(f"Cost of replacing pipes over unit headloss threshold: € {pipe_df['Replacement_Cost'].sum():,.2f}")
    print(pipe_df.dropna(subset=['Replacement_Cost']))
    print(f'Pipes which exceed headloss >{max_headloss}', critical_pipes)
    print(f'Total Investment Cost for Pump, Replaced Pipes and Tank: € {round(total_inv_cost,2):,}')
    
    # Individual nodal outputs to Excel
    wb = load_workbook(filename = f'{INP_file_name}_results/{INP_file_name}_results.xlsx')
    writer = pd.ExcelWriter(f'{INP_file_name}_results/{INP_file_name}_results.xlsx', engine='openpyxl')
    writer.book = wb
    writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)
        
    pipe_df_xl = pipe_df.dropna(subset=['Replacement_Cost'])
    pipe_df_xl.to_excel(writer, sheet_name=j, startrow=16)
    ws = wb[j]
        
    ws['A1'] = f'This sheet contains the results for tank connected to node: {j}'
    ws['A3'] = 'Pumping Parameters'
    ws['A4'] = 'Duty Head' ; ws['B4'] = round(Duty_Head, 2) ; ws['C4'] = 'm'
    ws['A5'] = 'Duty Flow' ; ws['B5'] = round(Duty_Flow*1000, 2) ; ws['C5'] = 'L/s'
    ws['A6'] = 'Actual average pumped flow:' ; ws['B6'] = round(pump_flows.mean(), 2) ; ws['C6'] = 'L/s'
    ws['A7'] = 'Cost:' ; ws['B7'] = f'€ {pump_cost[0:24].sum()[0]*3600:,.2f}' ; ws['C7'] = 'Euro per day'
    ws['A8'] = 'Energy:' ; ws['B8'] = round(pump_energy[0:24].sum()[0]/1000, 2) ; ws['C8'] = 'kWh/day'
    
    ws['A10'] = 'Balancing Tank Parameters'
    ws['A11'] = 'Elevation' ; ws['B11'] = round(tank_elev, 2) ; ws['C11'] = 'm above sea level'
    ws['A12'] = 'Tank height above ground' ; ws['B12'] = round(tank_height, 2) ; ws['C12'] = 'm above nearest node'
    ws['A13'] = 'Tank volume' ; ws['B13'] = round(tank_volume, 2) ; ws['C13'] = 'm3'
    
    ws['E3'] = 'Network Critical Results'
    ws['E4'] = 'Minimum Pressure' ; ws['F4'] = round(min_p, 2) ; ws['G4'] = 'mwc'
    ws['E5'] = 'Critical Hour' ; ws['F5'] = results.node['pressure'].loc[:, str(results.node['pressure'].loc[:, junctions].min().idxmin())].idxmin()/3600 ; ws['G5'] = 'hrs'
    ws['E6'] = 'Critical Node' ; ws['F6'] = results.node['pressure'].loc[:, junctions].min().idxmin()
    ws['E7'] = 'Critical Pipes' ; ws['F7'] = str(critical_pipes) ; ws['G7'] = 'Unit headloss >10m/km'
    
    ws['E10'] = 'Investment Cost Summary'
    ws['E11'] = 'Pump Cost' ; ws['F11'] = f'€ {pump_inv_cost:,.2f}'
    ws['E12'] = 'Tank Cost' ; ws['F12'] = f'€ {tank_inv_cost:,.2f}'
    ws['E13'] = 'Total Pipe Replacement Cost' ; ws['F13'] = f"€ {pipe_df['Replacement_Cost'].sum():,.2f}" 
    ws['E14'] = 'Total Investment Cost' ; ws['F14'] = f'€ {total_inv_cost:,.2f}'
    
    ws['A16'] = 'Table of individual pipes to be replaced'
       
    wb.save(f'{INP_file_name}_results/{INP_file_name}_results.xlsx')
    
    # Summary page output to excel
    node_list.append(j)
    Duty_Head_list.append(round(Duty_Head,2))
    Duty_Flow_list.append(round(Duty_Flow*1000,2))
    ActAvgFlow_list.append(round(pump_flows.mean(),2))
    Cost_list.append(f'€ {pump_cost[0:24].sum()[0]*3600:,.2f}')
    Energy_list.append(round(pump_energy[0:24].sum()[0]/1000,2))
    Tank_elev_list.append(round(tank_elev,2))
    Tank_height_list.append(round(tank_height, 2))
    Tank_volume_list.append(round(tank_volume, 2))
    Min_p_list.append(round(min_p, 2))
    Critical_hr_list.append(results.node['pressure'].loc[:, str(results.node['pressure'].loc[:, junctions].min().idxmin())].idxmin()/3600)
    Critical_node_list.append(results.node['pressure'].loc[:, junctions].min().idxmin())
    Critical_pipes_list.append(str(critical_pipes))
    Pump_inv_cost_list.append(f'€ {pump_inv_cost:,.2f}')
    Tank_inv_cost_list.append(f'€ {tank_inv_cost:,.2f}')
    Pipe_replacement_cost_list.append(f"€ {pipe_df['Replacement_Cost'].sum():,.2f}")
    Total_inv_cost_list.append(f'€ {total_inv_cost:,.2f}')
    Maintenance_list.append(f'€ {maintenance:,.2f}')  
    Annuity_list.append(f'€ {Annuity:,.2f}')
    Total_annual_exp_list.append(f'€ {Total_annual_exp:,.2f}')
    
    
    # Write EPANET.inp file
    wn.write_inpfile(f'{INP_file_name}_networks/{INP_file_name} tank_con_node {j}.inp', version=2.2)
    
    # Resetting Pipe DataFrame
    pipe_df = pipe_df.drop(["New_Diameter\n(mm)", "Supply_and_Installed_Cost\n($ USD/m)", 'Replacement_Cost'], 1)
    pipe_df = pipe_df.reset_index()
    # print(pipe_df)
    
wb = load_workbook(filename = f'{INP_file_name}_results/{INP_file_name}_results.xlsx')
writer = pd.ExcelWriter(f'{INP_file_name}_results/{INP_file_name}_results.xlsx', engine='openpyxl')
writer.book = wb
writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)

# Creating Summary DataFrame from summary dict{} 
summary = {"Node": node_list,
              "Max Water Age\n(h)": water_age_rounded,
              "Duty Head\n(m)": Duty_Head_list,
              "Duty Flow\n(L/s)": Duty_Flow_list,
              "Actual average pumped flow\n(L/s)": ActAvgFlow_list,
              "Cost\n(€ /day)": Cost_list,
              "Energy\n(kWh/day)": Energy_list,
              "Tank Elevation\n(metres above sea level)": Tank_elev_list,
              "Tank height above ground\n(metres above nearest node)": Tank_height_list,
              "Tank Volume\n(m3)": Tank_volume_list,
              "Minimum Pressure\n(mwc)": Min_p_list,
              "Critical Hour\n(hrs)": Critical_hr_list,
              "Critical Node": Critical_node_list,
              "Critical Pipes\n(list)": Critical_pipes_list,
              "Pump Investment Cost\n(Capital Investment)": Pump_inv_cost_list,
              "Tank Investment Cost\n(Capital Investment)": Tank_inv_cost_list,
              "Total Pipe Replacement Cost\n(Capital Investment)": Pipe_replacement_cost_list,
              "Total Investment Cost\n(Grand Total)": Total_inv_cost_list,
              "Maintenance Cost\n(pa)": Maintenance_list,
              "Annuity\n(pa)": Annuity_list,
              "Total Annual Expenditure": Total_annual_exp_list
        }
summary_df = pd.DataFrame.from_dict(summary)
summary_df.to_excel(writer, sheet_name='Summary', startrow=3)
wb.save(f'{INP_file_name}_results/{INP_file_name}_results.xlsx')





    
    
    
    

    
    
    
    
    
    
    
    
    

